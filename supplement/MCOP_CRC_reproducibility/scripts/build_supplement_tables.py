from __future__ import annotations

import json
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[3]
SUPP = ROOT / "supplement" / "MCOP_CRC_reproducibility"
OUT = ROOT / "outputs"
TABLE_DIR = SUPP / "supplementary"


def read_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(path)
    return pd.read_csv(path)


def save_table(df: pd.DataFrame, name: str) -> None:
    TABLE_DIR.mkdir(parents=True, exist_ok=True)
    df.to_csv(TABLE_DIR / f"{name}.csv", index=False, encoding="utf-8-sig")


def load_source(relative: str) -> pd.DataFrame:
    return read_csv(SUPP / "results" / relative)


def build_tables() -> dict[str, pd.DataFrame]:
    audit = load_source("phase2_primary/mcop_crc_phase2_audit.csv").iloc[0]
    primary = load_source("phase2h_survey/mcop_crc_phase2h_primary_reanalysis.csv")
    survey_compare = load_source("phase2h_survey/mcop_crc_phase2h_python_vs_standard_survey.csv")
    screen = read_csv(OUT / "environmental_crc_systematic_human_screen_fdr_v2.csv")
    robustness = read_csv(OUT / "environmental_crc_15axis_robustness_summary.csv")
    sensitivity = load_source("phase2h_survey/mcop_crc_phase2h_sensitivity_reanalysis.csv")
    per_cycle = read_csv(OUT / "mcop_crc_phase2_per_cycle.csv")
    heterogeneity = read_csv(OUT / "mcop_crc_phase2_cycle_interaction.csv")
    spline = read_csv(OUT / "mcop_crc_phase2_spline.csv")
    cycle_summary = read_csv(OUT / "mcop_crc_phase2_cycle_heterogeneity_summary.csv")
    assay_lod = read_csv(OUT / "mcop_crc_phase2_assay_lod_audit.csv")
    actionability = read_csv(OUT / "environmental_crc_267_actionability_matrix_v2.csv")
    actionability_flow = read_csv(OUT / "environmental_crc_267_actionability_flow.csv")
    figure5_lock = read_csv(OUT / "figure5_evidence_tier_lock.csv")
    bridge = read_csv(OUT / "mcop_phase2g_bridge_evidence_table.csv")
    phase2g_qc = json.loads((OUT / "mcop_phase2g_local_completion_audit.json").read_text(encoding="utf-8"))

    s1 = pd.DataFrame(
        [
            ["Harmonized NHANES records", int(audit["harmonized_rows"]), "Seven-cycle harmonized frame; adults and core covariates available in source construction"],
            ["MCOP available records", int(audit["mcop_available_rows"]), "URXCOP available in 2005–06 through 2017–18"],
            ["Primary exposure/outcome records", int(audit["primary_exposure_outcome_rows"]), "CRC case definition and cancer-free control frame before complete-case covariate restriction"],
            ["Primary CRC cases before complete-case restriction", int(audit["primary_crc_cases"]), "Current/previous CRC definition used by frozen NHANES analysis"],
            ["Cancer-free controls before complete-case restriction", int(audit["primary_cancer_free_controls"]), "Reference group for the primary CRC-versus-cancer-free analysis"],
            ["MCOP above LOD", int(audit["primary_mcop_above_lod_n"]), "Detected MCOP values in the primary exposure/outcome frame"],
            ["Primary complete-case analytic sample", 9936, "All frozen covariates, survey design variables and MCOP available"],
            ["Primary complete-case CRC cases", 70, "Event count used in the frozen primary model"],
            ["Primary complete-case controls", 9866, "Cancer-free controls used in the frozen primary model"],
        ],
        columns=["Stage", "N", "Definition / audit note"],
    )

    s2 = primary[["Analysis", "Cycles", "status", "N", "CRC_N", "Control_N", "OR", "CI_low", "CI_high", "P", "design_df", "PSU_N", "strata_N", "message"]].copy()
    s2.insert(0, "Result_family", "Primary model / weight audit")
    compare = survey_compare.copy()
    compare.insert(0, "Result_family", "Independent Python vs R survey implementation")
    s2 = pd.concat([s2, compare], ignore_index=True, sort=False)

    s3 = screen.copy().sort_values("screen_rank")
    s3 = s3[["screen_rank", "axis_key", "exposure_axis", "primary_biomarker", "biological_matrix", "eligible_chemical_count", "n_cycles_available", "fit_N", "fit_crc_cases", "OR", "CI_low", "CI_high", "P", "BH_FDR", "status", "message"]]

    s4 = robustness[["axis_key", "exposure_axis", "primary_biomarker", "eligible_chemical_count", "primary_OR", "primary_CI_low", "primary_CI_high", "primary_P", "primary_BH_FDR_15axis", "robustness_tier", "fit_status", "fit_N", "fit_crc_cases", "F_note", "L_note", "C_note", "H_note", "D_note", "T_note", "A_note", "E_note", "timing_max_abs_logOR_delta", "tail_max_abs_logOR_delta"]].copy()

    keep = ["Analysis", "analysis_family", "Dropped_cycle", "Sex_group", "N", "CRC_N", "Control_N", "OR", "CI_low", "CI_high", "P", "MCOP_BH_FDR", "Secondary_exposure", "Secondary_OR", "Secondary_CI_low", "Secondary_CI_high", "Secondary_P", "Excluded_known_timing_CRC_N", "Exposure", "Excluded_fraction", "Normalization", "Burden_definition"]
    s5 = sensitivity[[c for c in keep if c in sensitivity.columns]].copy()

    s6_cycle = per_cycle[["Cycle", "N", "CRC_N", "OR", "CI_low", "CI_high", "P", "status", "design_df", "PSU_N", "strata_N"]].copy()
    s6_cycle.insert(0, "Result_family", "Cycle-specific model")
    s6_int = heterogeneity.copy()
    s6_int.insert(0, "Result_family", "MCOP × cycle interaction")
    s6_spline = spline.copy()
    s6_spline.insert(0, "Result_family", "Restricted cubic spline")
    s6 = pd.concat([s6_cycle, s6_int, s6_spline], ignore_index=True, sort=False)

    s7_cycle = cycle_summary[["cycle", "Model_complete_case_N", "Model_complete_case_CRC_cases", "Model_CRC_MCOP_median_ng_mL", "Model_control_MCOP_median_ng_mL", "Model_case_control_MCOP_median_ratio", "MCOP_above_LOD_pct", "LLOD_ng_mL_codebook", "MCOP_median_ng_mL", "MCOP_P95_ng_mL", "MCOP_P99_ng_mL", "assay_platform"]].copy()
    s7_cycle.insert(0, "Result_family", "Cycle-level exposure and assay audit")
    s7_lod = assay_lod.copy()
    s7_lod.insert(0, "Result_family", "Analyte LOD audit")
    s7 = pd.concat([s7_cycle, s7_lod], ignore_index=True, sort=False)

    s8 = actionability.copy()
    s8_flow = actionability_flow.copy()
    s8_flow.insert(0, "Table_section", "267-chemical attrition flow")
    s8.insert(0, "Table_section", "267-chemical actionability matrix")
    s8 = pd.concat([s8_flow, s8], ignore_index=True, sort=False)

    qc_rows = pd.DataFrame(
        [
            ["Phase2G local H5AD equivalence", "source_h5ad_equivalence_qc", phase2g_qc.get("source_h5ad_equivalence_qc"), "Official H5AD source used for final local Phase2G analysis"],
            ["Phase2G local H5AD equivalence", "eligible_cells", phase2g_qc.get("eligible_cells"), "Eligible epithelial cells"],
            ["Phase2G local H5AD equivalence", "paired_donors", phase2g_qc.get("paired_donors"), "Paired donors in frozen epithelial analysis"],
            ["Phase2G local H5AD equivalence", "frozen_ppar_delta", phase2g_qc.get("frozen_ppar_delta"), "Paired donor PPAR/NR median delta"],
            ["Phase2G local H5AD equivalence", "frozen_ppar_p", phase2g_qc.get("frozen_ppar_p"), "Paired donor PPAR/NR p-value"],
        ],
        columns=["Result_family", "Metric", "Value", "Interpretation"],
    )
    s9_lock = figure5_lock.copy()
    s9_lock.insert(0, "Table_section", "Figure 5 evidence tier lock")
    bridge2 = bridge.copy()
    bridge2.insert(0, "Table_section", "DINP-axis molecular bridge evidence")
    s9_qc = qc_rows.copy()
    s9_qc.insert(0, "Table_section", "Phase2G source QC")
    s9 = pd.concat([s9_lock, bridge2, s9_qc], ignore_index=True, sort=False)

    return {"S1_Sample_selection": s1, "S2_Primary_model": s2, "S3_15axis_screen": s3, "S4_Robustness": s4, "S5_Sensitivity": s5, "S6_Heterogeneity_shape": s6, "S7_Exposure_LOD": s7, "S8_267_actionability": s8, "S9_Transcriptomic_evidence": s9}


def write_workbook(tables: dict[str, pd.DataFrame]) -> None:
    TABLE_DIR.mkdir(parents=True, exist_ok=True)
    path = TABLE_DIR / "MCOP_CRC_Supplement_Tables.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for name, df in tables.items():
        ws = wb.create_sheet(name[:31])
        for j, col in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=j, value=str(col))
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in df.itertuples(index=False, name=None):
            ws.append([None if pd.isna(v) else v for v in row])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col_idx, col in enumerate(df.columns, 1):
            sample = [str(col)] + [str(v) for v in df.iloc[:100, col_idx - 1].tolist()]
            width = min(42, max(10, max(len(x) for x in sample) + 2))
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)


def main() -> None:
    tables = build_tables()
    for name, df in tables.items():
        save_table(df, name)
    write_workbook(tables)
    print(json.dumps({k: list(v.shape) for k, v in tables.items()}, ensure_ascii=False))


if __name__ == "__main__":
    main()
