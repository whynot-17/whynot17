from __future__ import annotations

import hashlib
import re
import zipfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[3]
OUT = ROOT / "outputs"
PKG = ROOT / "supplement" / "MCOP_CRC_submission"
QA = PKG / "QA"
QA.mkdir(parents=True, exist_ok=True)


def ok(name, condition, observed, expected):
    return {"check": name, "status": "PASS" if condition else "FAIL", "observed": str(observed), "expected": str(expected)}


def docx_text(path):
    with zipfile.ZipFile(path) as z:
        xml = z.read("word/document.xml").decode("utf-8")
    return re.sub(r"<[^>]+>", " ", xml)


def sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


flow = pd.read_csv(OUT / "environmental_crc_267_actionability_flow.csv")
axes = pd.read_csv(OUT / "environmental_crc_267_human_testable_candidates.csv")
screen = pd.read_csv(OUT / "environmental_crc_systematic_human_screen_fdr_v2.csv")
score = pd.read_csv(OUT / "environmental_crc_15axis_robustness_scorecard.csv")
compare = pd.read_csv(OUT / "mcop_crc_phase2h_python_vs_standard_survey.csv").iloc[0]
defs = pd.read_csv(OUT / "figure5_ppar_definition_comparison.csv")

counts = dict(zip(flow.stage, flow.n))
checks = [
    ok("Starting chemical universe", counts.get("total_core_chemicals") == 267, counts.get("total_core_chemicals"), 267),
    ok("Human-testable mappings", counts.get("E_X_B_D_C_and_T_testable") == 87, counts.get("E_X_B_D_C_and_T_testable"), 87),
    ok("Strict-eligible mappings", counts.get("strict_eligibility") == 27, counts.get("strict_eligibility"), 27),
    ok("Unique biomarker tests", len(axes) == 15, len(axes), 15),
    ok("Unified screen tests", len(screen) == 15, len(screen), 15),
    ok("BH-FDR-supported tests", int((screen.BH_FDR < 0.05).sum()) == 2, int((screen.BH_FDR < 0.05).sum()), 2),
    ok("Robust Tier A axes", int((score.robustness_tier == "Robust Tier A").sum()) == 1, int((score.robustness_tier == "Robust Tier A").sum()), 1),
    ok("Primary complete-case N", int(compare.r_N) == 9936, int(compare.r_N), 9936),
    ok("Primary CRC events", int(compare.r_CRC_N) == 70, int(compare.r_CRC_N), 70),
    ok("Primary OR", abs(compare.r_OR - 1.2455068) < 1e-6, f"{compare.r_OR:.7f}", "1.2455068 ±1e-6"),
    ok("Primary design-df P", abs(compare.r_P_design_df - 0.0033113) < 1e-6, f"{compare.r_P_design_df:.7f}", "0.0033113 ±1e-6"),
    ok("R/Python logOR agreement", compare.absolute_logOR_difference < 1e-10, f"{compare.absolute_logOR_difference:.3e}", "<1e-10"),
]

core = defs.loc[defs.definition.eq("Frozen 7-gene PPAR/NR core")].iloc[0]
checks += [
    ok("Frozen PPAR/NR paired donors", int(core.n_paired_donors) == 36, int(core.n_paired_donors), 36),
    ok("Frozen PPAR/NR median delta", abs(core.median_delta + 0.418601) < 1e-6, f"{core.median_delta:.6f}", "-0.418601 ±1e-6"),
    ok("Frozen PPAR/NR FDR", abs(core.BH_FDR - 9.299838e-7) < 1e-12, f"{core.BH_FDR:.6e}", "9.299838e-7 ±1e-12"),
]

supp_docx = PKG / "MCOP_CRC_Supplementary_Information.docx"
supp_text = docx_text(supp_docx)
checks += [
    ok("Supplement avoids causal language", "DINP causes CRC" not in supp_text and "MCOP causes CRC" not in supp_text, "no prohibited direct-causal phrase", "no prohibited direct-causal phrase"),
    ok("Supplement locks chemical identity", all(x in supp_text for x in ["MiNP", "DINP", "MCOP", "chemical equivalence"]), "identity lock present", "identity lock present"),
]

tables = load_workbook(PKG / "MCOP_CRC_Supplement_Tables.xlsx", read_only=True)
source = load_workbook(PKG / "MCOP_CRC_Source_Data.xlsx", read_only=True)
checks += [
    ok("Supplementary table sheets", len(tables.sheetnames) == 9, len(tables.sheetnames), 9),
    ok("Figure source-data sheets", len(source.sheetnames) >= 12, len(source.sheetnames), ">=12"),
]

for i in range(1, 5):
    stem = ["Figure_S1_actionability_audit", "Figure_S2_human_screen_robustness", "Figure_S3_cycle_exposure_audit", "Figure_S4_ppar_state_evidence"][i-1]
    checks.append(ok(f"Figure S{i} export triad", all((PKG / "figures" / f"{stem}.{ext}").exists() for ext in ["pdf","svg","png"]), "PDF/SVG/PNG", "PDF/SVG/PNG"))

audit = pd.DataFrame(checks)
audit.to_csv(QA / "submission_numeric_consistency_audit.csv", index=False)

desktop_ms = Path(r"C:\Users\21634\Desktop\whynot.docx")
integration = []
if desktop_ms.exists():
    mt = docx_text(desktop_ms)
    n_placeholder = len(re.findall(r"Supplementary\s+Table\s+X", mt, flags=re.I))
    integration.append(("Supplementary Table X placeholders in current manuscript", n_placeholder,
                        "Replace with S8 (gate fields), S1/S8 (attrition), and S4 (robustness rubric)."))
integration.append(("Main Figure 2 asset", "requires final integration check",
                    "Repository still contains pre-lock candidate-triage Figure2 files; use the outcome-blinded 267→87→15 architecture before submission."))

manifest = []
for p in sorted(PKG.rglob("*")):
    if (p.is_file() and "rendered" not in p.parts and p.suffix.lower() not in {".html"}
            and p.name != "submission_file_manifest_sha256.csv"):
        manifest.append({"relative_path": str(p.relative_to(PKG)), "bytes": p.stat().st_size, "sha256": sha256(p)})
pd.DataFrame(manifest).to_csv(QA / "submission_file_manifest_sha256.csv", index=False)

lines = ["# MCOP–CRC Supplement consistency and submission audit", "",
         f"Numeric/package checks: **{(audit.status=='PASS').sum()} PASS, {(audit.status=='FAIL').sum()} FAIL**.", "",
         "## Frozen numeric consistency", "", "| Check | Status | Observed | Expected |", "|---|---:|---:|---:|"]
for r in audit.itertuples(index=False):
    lines.append(f"| {r.check} | {r.status} | {r.observed} | {r.expected} |")
lines += ["", "## Main-manuscript integration items", ""]
for name, obs, action in integration:
    lines.append(f"- **{name}:** {obs}. {action}")
lines += ["", "## Interpretation lock", "",
          "- NHANES is cross-sectional and estimates prevalent CRC association, not incident risk.",
          "- MiNP, DINP parent and MCOP are chemically distinct; MCOP is the urinary biomarker for a DINP-related axis.",
          "- PPAR/NR is an independently observed CRC epithelial disease-state program; the DINP/MCOP-to-state bridge remains untested.",
          "- LOCO analyses are overlapping pooled re-estimates, not seven independent replications.",
          "- The incomplete 35/36 live Census staged cache remains an INFO item; final transcriptomic inference uses the validated official H5AD."]
(QA / "supplement_consistency_report.md").write_text("\n".join(lines), encoding="utf-8")
print(audit.status.value_counts().to_dict())
